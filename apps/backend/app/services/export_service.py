"""
Export service - generates CSV and Excel files for orders and invoices.
"""

import csv
import io
import logging

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.core.timezone import utc_to_local
from app.models.invoice import Invoice
from app.models.order import Order

logger = logging.getLogger(__name__)

# Invoice type mapping
INVOICE_TYPE_NAMES = {
    "01": "Factura",
    "03": "Boleta",
    "07": "Nota de Crédito",
    "08": "Nota de Débito",
}

EFACT_STATUS_NAMES = {
    "pending": "Pendiente",
    "processing": "Procesando",
    "success": "Validado",
    "error": "Error",
}

# Column definitions
ORDER_COLUMNS = [
    "ID",
    "ID Draft Shopify",
    "ID Orden",
    "Cliente",
    "Email",
    "Estado",
    "Monto",
    "Moneda",
    "Canal",
    "Método de Pago",
    "Fecha de Creación",
    "Fecha de Validación",
]

INVOICE_COLUMNS = [
    "Serie-Número",
    "Tipo",
    "Cliente",
    "Documento",
    "Subtotal",
    "IGV",
    "Total",
    "Moneda",
    "Estado SUNAT",
    "Fecha de Emisión",
]


def _format_local_datetime(dt, tz_name: str) -> str:
    """Format a UTC datetime as local datetime string."""
    if dt is None:
        return ""
    local_dt = utc_to_local(dt, tz_name)
    return local_dt.strftime("%d/%m/%Y %H:%M")


def _get_order_source_id(order: Order) -> str:
    """Get display ID for order source platform."""
    if order.shopify_order_id:
        # Extract numeric part from gid://shopify/Order/123
        parts = order.shopify_order_id.split("/")
        return parts[-1] if parts else order.shopify_order_id
    if order.woocommerce_order_id:
        return str(order.woocommerce_order_id)
    return ""


class ExportService:
    """Service for generating CSV and Excel export files."""

    def export_orders_csv(self, orders: list[Order], tz_name: str) -> io.BytesIO:
        """Generate CSV file with order data."""
        output = io.BytesIO()
        # UTF-8 BOM for Excel compatibility
        output.write(b"\xef\xbb\xbf")
        wrapper = io.TextIOWrapper(output, encoding="utf-8", newline="")

        writer = csv.writer(wrapper)
        writer.writerow(ORDER_COLUMNS)

        for order in orders:
            writer.writerow([
                order.id,
                order.shopify_draft_order_id or "",
                _get_order_source_id(order),
                order.customer_name or "",
                order.customer_email,
                order.status or "",
                f"{order.total_price:.2f}",
                order.currency,
                order.channel or "",
                order.payment_method or "",
                _format_local_datetime(order.created_at, tz_name),
                _format_local_datetime(order.validated_at, tz_name),
            ])

        wrapper.detach()
        output.seek(0)
        return output

    def export_orders_excel(self, orders: list[Order], tz_name: str) -> io.BytesIO:
        """Generate Excel file with order data."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Pedidos"

        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")

        for col_idx, col_name in enumerate(ORDER_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row_idx, order in enumerate(orders, 2):
            ws.cell(row=row_idx, column=1, value=order.id)
            ws.cell(row=row_idx, column=2, value=order.shopify_draft_order_id or "")
            ws.cell(row=row_idx, column=3, value=_get_order_source_id(order))
            ws.cell(row=row_idx, column=4, value=order.customer_name or "")
            ws.cell(row=row_idx, column=5, value=order.customer_email)
            ws.cell(row=row_idx, column=6, value=order.status or "")
            ws.cell(row=row_idx, column=7, value=round(order.total_price, 2))
            ws.cell(row=row_idx, column=8, value=order.currency)
            ws.cell(row=row_idx, column=9, value=order.channel or "")
            ws.cell(row=row_idx, column=10, value=order.payment_method or "")
            ws.cell(row=row_idx, column=11, value=_format_local_datetime(order.created_at, tz_name))
            ws.cell(row=row_idx, column=12, value=_format_local_datetime(order.validated_at, tz_name))

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def export_invoices_csv(self, invoices: list[Invoice], tz_name: str) -> io.BytesIO:
        """Generate CSV file with invoice data."""
        output = io.BytesIO()
        output.write(b"\xef\xbb\xbf")
        wrapper = io.TextIOWrapper(output, encoding="utf-8", newline="")

        writer = csv.writer(wrapper)
        writer.writerow(INVOICE_COLUMNS)

        for inv in invoices:
            full_number = f"{inv.serie}-{inv.correlativo:08d}"
            writer.writerow([
                full_number,
                INVOICE_TYPE_NAMES.get(inv.invoice_type, inv.invoice_type),
                inv.cliente_razon_social,
                inv.cliente_numero_documento,
                f"{inv.subtotal:.2f}",
                f"{inv.igv:.2f}",
                f"{inv.total:.2f}",
                inv.currency,
                EFACT_STATUS_NAMES.get(inv.efact_status, inv.efact_status),
                _format_local_datetime(inv.created_at, tz_name),
            ])

        wrapper.detach()
        output.seek(0)
        return output

    def export_invoices_excel(self, invoices: list[Invoice], tz_name: str) -> io.BytesIO:
        """Generate Excel file with invoice data."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Comprobantes"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")

        for col_idx, col_name in enumerate(INVOICE_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row_idx, inv in enumerate(invoices, 2):
            full_number = f"{inv.serie}-{inv.correlativo:08d}"
            ws.cell(row=row_idx, column=1, value=full_number)
            ws.cell(row=row_idx, column=2, value=INVOICE_TYPE_NAMES.get(inv.invoice_type, inv.invoice_type))
            ws.cell(row=row_idx, column=3, value=inv.cliente_razon_social)
            ws.cell(row=row_idx, column=4, value=inv.cliente_numero_documento)
            ws.cell(row=row_idx, column=5, value=round(inv.subtotal, 2))
            ws.cell(row=row_idx, column=6, value=round(inv.igv, 2))
            ws.cell(row=row_idx, column=7, value=round(inv.total, 2))
            ws.cell(row=row_idx, column=8, value=inv.currency)
            ws.cell(row=row_idx, column=9, value=EFACT_STATUS_NAMES.get(inv.efact_status, inv.efact_status))
            ws.cell(row=row_idx, column=10, value=_format_local_datetime(inv.created_at, tz_name))

        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output


export_service = ExportService()

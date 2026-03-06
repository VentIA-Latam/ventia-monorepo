"""
Tests for app/services/export_service.py — CSV/Excel export generation.
"""

import csv
import io
from datetime import datetime
from unittest.mock import MagicMock

import pytest
from openpyxl import load_workbook

from app.models.invoice import Invoice
from app.models.order import Order
from app.services.export_service import (
    EFACT_STATUS_NAMES,
    INVOICE_COLUMNS,
    INVOICE_TYPE_NAMES,
    ORDER_COLUMNS,
    ExportService,
    _format_local_datetime,
    _get_order_source_id,
)

TZ = "America/Lima"


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Fixtures
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━


@pytest.fixture
def service() -> ExportService:
    return ExportService()


@pytest.fixture
def mock_orders() -> list[MagicMock]:
    """Two mock orders: one Shopify, one WooCommerce."""
    o1 = MagicMock(spec=Order)
    o1.id = 1
    o1.shopify_draft_order_id = "gid://shopify/DraftOrder/999"
    o1.shopify_order_id = "gid://shopify/Order/12345"
    o1.woocommerce_order_id = None
    o1.customer_name = "Juan Perez"
    o1.customer_email = "juan@example.com"
    o1.status = "Pagado"
    o1.total_price = 118.0
    o1.currency = "PEN"
    o1.channel = "shopify"
    o1.payment_method = "Transferencia"
    o1.created_at = datetime(2024, 1, 15, 5, 0, 0)  # 00:00 Lima
    o1.validated_at = datetime(2024, 1, 15, 8, 30, 0)

    o2 = MagicMock(spec=Order)
    o2.id = 2
    o2.shopify_draft_order_id = None
    o2.shopify_order_id = None
    o2.woocommerce_order_id = 456
    o2.customer_name = "Maria Garcia"
    o2.customer_email = "maria@example.com"
    o2.status = "Pendiente"
    o2.total_price = 236.50
    o2.currency = "USD"
    o2.channel = "woocommerce"
    o2.payment_method = None
    o2.created_at = datetime(2024, 2, 20, 15, 0, 0)
    o2.validated_at = None

    return [o1, o2]


@pytest.fixture
def mock_invoices() -> list[MagicMock]:
    """Two mock invoices: one Factura, one Boleta."""
    inv1 = MagicMock(spec=Invoice)
    inv1.serie = "F001"
    inv1.correlativo = 1
    inv1.invoice_type = "01"
    inv1.cliente_razon_social = "Empresa SAC"
    inv1.cliente_numero_documento = "20123456789"
    inv1.subtotal = 100.00
    inv1.igv = 18.00
    inv1.total = 118.00
    inv1.currency = "PEN"
    inv1.efact_status = "success"
    inv1.created_at = datetime(2024, 3, 10, 5, 0, 0)

    inv2 = MagicMock(spec=Invoice)
    inv2.serie = "B001"
    inv2.correlativo = 42
    inv2.invoice_type = "03"
    inv2.cliente_razon_social = "Juan Perez"
    inv2.cliente_numero_documento = "12345678"
    inv2.subtotal = 200.00
    inv2.igv = 36.00
    inv2.total = 236.00
    inv2.currency = "PEN"
    inv2.efact_status = "pending"
    inv2.created_at = datetime(2024, 3, 11, 12, 0, 0)

    return [inv1, inv2]


def _parse_csv(output: io.BytesIO) -> list[list[str]]:
    """Helper: read CSV from BytesIO, skip BOM."""
    raw = output.read()
    # Strip UTF-8 BOM
    if raw.startswith(b"\xef\xbb\xbf"):
        raw = raw[3:]
    text = raw.decode("utf-8")
    reader = csv.reader(io.StringIO(text))
    return list(reader)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Export Orders CSV
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━


class TestExportOrdersCsv:
    def test_csv_starts_with_utf8_bom(self, service, mock_orders):
        output = service.export_orders_csv(mock_orders, TZ)
        first_bytes = output.read(3)
        assert first_bytes == b"\xef\xbb\xbf"

    def test_csv_header_row_matches_columns(self, service, mock_orders):
        output = service.export_orders_csv(mock_orders, TZ)
        rows = _parse_csv(output)
        assert rows[0] == ORDER_COLUMNS

    def test_csv_row_count_matches_orders(self, service, mock_orders):
        output = service.export_orders_csv(mock_orders, TZ)
        rows = _parse_csv(output)
        assert len(rows) == 1 + len(mock_orders)  # header + data

    def test_csv_shopify_gid_extracted(self, service, mock_orders):
        """gid://shopify/Order/12345 → '12345'."""
        output = service.export_orders_csv(mock_orders, TZ)
        rows = _parse_csv(output)
        # Row 1 (first order), column index 2 = "ID Orden"
        assert rows[1][2] == "12345"

    def test_csv_woocommerce_id_as_string(self, service, mock_orders):
        """WooCommerce int ID → string representation."""
        output = service.export_orders_csv(mock_orders, TZ)
        rows = _parse_csv(output)
        # Row 2 (second order), column index 2 = "ID Orden"
        assert rows[2][2] == "456"

    def test_csv_dates_converted_to_timezone(self, service, mock_orders):
        """UTC dates converted to Lima timezone in dd/mm/YYYY HH:MM format."""
        output = service.export_orders_csv(mock_orders, TZ)
        rows = _parse_csv(output)
        # First order: created_at=2024-01-15 05:00 UTC → 00:00 Lima
        assert rows[1][10] == "15/01/2024 00:00"

    def test_csv_none_validated_at_empty(self, service, mock_orders):
        """None validated_at → empty string."""
        output = service.export_orders_csv(mock_orders, TZ)
        rows = _parse_csv(output)
        # Second order has validated_at=None
        assert rows[2][11] == ""

    def test_csv_total_price_two_decimals(self, service, mock_orders):
        """Total price formatted with 2 decimal places."""
        output = service.export_orders_csv(mock_orders, TZ)
        rows = _parse_csv(output)
        assert rows[1][6] == "118.00"
        assert rows[2][6] == "236.50"

    def test_csv_empty_orders_header_only(self, service):
        output = service.export_orders_csv([], TZ)
        rows = _parse_csv(output)
        assert len(rows) == 1
        assert rows[0] == ORDER_COLUMNS

    def test_csv_returns_seeked_bytesio(self, service, mock_orders):
        """BytesIO position should be at 0 after generation."""
        output = service.export_orders_csv(mock_orders, TZ)
        assert output.tell() == 0


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Export Orders Excel
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━


class TestExportOrdersExcel:
    def test_excel_valid_workbook(self, service, mock_orders):
        output = service.export_orders_excel(mock_orders, TZ)
        wb = load_workbook(output)
        assert wb is not None

    def test_excel_sheet_name_pedidos(self, service, mock_orders):
        output = service.export_orders_excel(mock_orders, TZ)
        wb = load_workbook(output)
        assert wb.active.title == "Pedidos"

    def test_excel_header_bold_white(self, service, mock_orders):
        output = service.export_orders_excel(mock_orders, TZ)
        wb = load_workbook(output)
        ws = wb.active
        header_cell = ws.cell(row=1, column=1)
        assert header_cell.font.bold is True
        assert header_cell.font.color.rgb == "00FFFFFF"

    def test_excel_data_row_count(self, service, mock_orders):
        output = service.export_orders_excel(mock_orders, TZ)
        wb = load_workbook(output)
        ws = wb.active
        # max_row = header + data rows
        assert ws.max_row == 1 + len(mock_orders)

    def test_excel_total_price_numeric(self, service, mock_orders):
        """Total price should be a number, not a string."""
        output = service.export_orders_excel(mock_orders, TZ)
        wb = load_workbook(output)
        ws = wb.active
        # Column 7 = "Monto", row 2 = first data row
        value = ws.cell(row=2, column=7).value
        assert isinstance(value, (int, float))
        assert value == 118.0

    def test_excel_empty_orders_header_only(self, service):
        output = service.export_orders_excel([], TZ)
        wb = load_workbook(output)
        ws = wb.active
        assert ws.max_row == 1


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Export Invoices CSV
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━


class TestExportInvoicesCsv:
    def test_csv_header_matches_invoice_columns(self, service, mock_invoices):
        output = service.export_invoices_csv(mock_invoices, TZ)
        rows = _parse_csv(output)
        assert rows[0] == INVOICE_COLUMNS

    def test_csv_serie_correlativo_8_digits(self, service, mock_invoices):
        """Correlativo formatted as 8 digits: B001-00000042."""
        output = service.export_invoices_csv(mock_invoices, TZ)
        rows = _parse_csv(output)
        assert rows[1][0] == "F001-00000001"
        assert rows[2][0] == "B001-00000042"

    def test_csv_invoice_type_mapped(self, service, mock_invoices):
        """'01' → 'Factura', '03' → 'Boleta'."""
        output = service.export_invoices_csv(mock_invoices, TZ)
        rows = _parse_csv(output)
        assert rows[1][1] == "Factura"
        assert rows[2][1] == "Boleta"

    def test_csv_efact_status_mapped(self, service, mock_invoices):
        """'success' → 'Validado', 'pending' → 'Pendiente'."""
        output = service.export_invoices_csv(mock_invoices, TZ)
        rows = _parse_csv(output)
        assert rows[1][8] == "Validado"
        assert rows[2][8] == "Pendiente"

    def test_csv_unknown_type_fallback(self, service):
        """Unknown invoice type shows raw value."""
        inv = MagicMock(spec=Invoice)
        inv.serie = "X001"
        inv.correlativo = 1
        inv.invoice_type = "99"
        inv.cliente_razon_social = "Test"
        inv.cliente_numero_documento = "000"
        inv.subtotal = 10.00
        inv.igv = 1.80
        inv.total = 11.80
        inv.currency = "PEN"
        inv.efact_status = "unknown_status"
        inv.created_at = datetime(2024, 1, 1, 5, 0, 0)

        output = service.export_invoices_csv([inv], TZ)
        rows = _parse_csv(output)
        assert rows[1][1] == "99"
        assert rows[1][8] == "unknown_status"

    def test_csv_amounts_two_decimals(self, service, mock_invoices):
        output = service.export_invoices_csv(mock_invoices, TZ)
        rows = _parse_csv(output)
        # First invoice: 100.00, 18.00, 118.00
        assert rows[1][4] == "100.00"
        assert rows[1][5] == "18.00"
        assert rows[1][6] == "118.00"


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Export Invoices Excel
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━


class TestExportInvoicesExcel:
    def test_excel_sheet_name_comprobantes(self, service, mock_invoices):
        output = service.export_invoices_excel(mock_invoices, TZ)
        wb = load_workbook(output)
        assert wb.active.title == "Comprobantes"

    def test_excel_numeric_amounts(self, service, mock_invoices):
        output = service.export_invoices_excel(mock_invoices, TZ)
        wb = load_workbook(output)
        ws = wb.active
        # Row 2, columns 5/6/7 = subtotal/igv/total
        assert isinstance(ws.cell(row=2, column=5).value, (int, float))
        assert ws.cell(row=2, column=5).value == 100.0
        assert ws.cell(row=2, column=6).value == 18.0
        assert ws.cell(row=2, column=7).value == 118.0

    def test_excel_data_row_count(self, service, mock_invoices):
        output = service.export_invoices_excel(mock_invoices, TZ)
        wb = load_workbook(output)
        ws = wb.active
        assert ws.max_row == 1 + len(mock_invoices)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Helper functions
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━


class TestHelpers:
    def test_format_none_returns_empty(self):
        assert _format_local_datetime(None, TZ) == ""

    def test_format_correct_pattern(self):
        result = _format_local_datetime(datetime(2024, 1, 15, 5, 0, 0), TZ)
        assert result == "15/01/2024 00:00"

    def test_shopify_gid_extracts_id(self):
        order = MagicMock(spec=Order)
        order.shopify_order_id = "gid://shopify/Order/12345"
        order.woocommerce_order_id = None
        assert _get_order_source_id(order) == "12345"

    def test_woocommerce_returns_string(self):
        order = MagicMock(spec=Order)
        order.shopify_order_id = None
        order.woocommerce_order_id = 789
        assert _get_order_source_id(order) == "789"

    def test_no_platform_id_empty(self):
        order = MagicMock(spec=Order)
        order.shopify_order_id = None
        order.woocommerce_order_id = None
        assert _get_order_source_id(order) == ""
